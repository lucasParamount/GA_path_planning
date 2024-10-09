import random
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as patches
import time
import openpyxl


class Env:
    def __init__(self):
        # 定义障碍物
        self.obs_rectangle = [(639, 726, 86, 70), (750, 726, 94, 68), (672, 826, 18, 74),
                              (750, 820, 80, 80), (850, 790, 128, 110), (1006, 724, 95, 176),
                              (1113, 784, 87, 116), (600, 985, 230, 185), (385, 895, 70, 100),
                              (385, 1013, 81, 60), (725, 521, 134, 183), (893, 497, 126, 207)]
        self.width = 1500
        self.height = 1200

    def is_in_bounds(self, point):
        """ Check if point is inside environment bounds """
        x, y = point
        return 0 <= x <= self.width and 0 <= y <= self.height

    def is_collision(self, point):
        """ Check if a point collides with any rectangle obstacles """
        x, y = point
        for (ox, oy, w, h) in self.obs_rectangle:
            if ox <= x <= ox + w and oy <= y <= oy + h:
                return True
        return False

    def is_collision_in_segment(self, p1, p2):
        """ Check if a line segment between two points collides with any obstacles """
        for (ox, oy, w, h) in self.obs_rectangle:
            if self.line_rect_collision(p1, p2, (ox, oy, w, h)):
                return True
        return False

    def line_rect_collision(self, p1, p2, rect):
        """ Check if a line segment between p1 and p2 intersects a rectangle """
        ox, oy, w, h = rect
        # Define the four corners of the rectangle
        corners = [(ox, oy), (ox + w, oy), (ox + w, oy + h), (ox, oy + h)]
        # Define the four edges of the rectangle
        edges = [(corners[0], corners[1]), (corners[1], corners[2]),
                 (corners[2], corners[3]), (corners[3], corners[0])]

        # Check if the line intersects any of the edges
        for edge_start, edge_end in edges:
            if self.line_intersects(p1, p2, edge_start, edge_end):
                return True
        return False

    def line_intersects(self, p1, p2, p3, p4):
        """ Check if two line segments (p1, p2) and (p3, p4) intersect """
        def ccw(A, B, C):
            return (C[1] - A[1]) * (B[0] - A[0]) > (B[1] - A[1]) * (C[0] - A[0])

        return ccw(p1, p3, p4) != ccw(p2, p3, p4) and ccw(p1, p2, p3) != ccw(p1, p2, p4)


class GeneticAlgorithmPathPlanner:
    def __init__(self, start, goal, env, population_size=50, generations=500, mutation_rate=0.1, n_intermediate=10):
        self.start = np.array(start)
        self.goal = np.array(goal)
        self.env = env
        self.population_size = population_size
        self.generations = generations
        self.mutation_rate = mutation_rate
        self.n_intermediate = n_intermediate
        self.population = self.init_population()

    def init_population(self):
        """ Initialize population with random paths """
        population = []
        for _ in range(self.population_size):
            path = self.generate_random_path()
            population.append(path)
        return population

    def generate_random_path(self):
        """ Generate a random path """
        intermediate_points = [self.random_point_in_free_space() for _ in range(self.n_intermediate)]
        return [self.start] + intermediate_points + [self.goal]

    def random_point_in_free_space(self):
        """ Generate a random point that is within the bounds and not in collision """
        while True:
            point = np.random.randint(0, [self.env.width, self.env.height])
            if not self.env.is_collision(point):
                return point

    def fitness(self, path):
        """ Calculate fitness based on path length and collision """
        total_length = 0
        for i in range(len(path) - 1):
            segment_length = np.linalg.norm(path[i + 1] - path[i])
            if self.env.is_collision_in_segment(path[i], path[i + 1]):
                return float('inf')  # If any segment collides, mark as invalid
            total_length += segment_length
        return total_length

    def selection(self):
        """ Select the top-performing half of the population """
        ranked_population = sorted(self.population, key=lambda path: self.fitness(path))
        return ranked_population[:self.population_size // 2]

    def crossover(self, parent1, parent2):
        """ Perform crossover by mixing the intermediate points of two parents """
        crossover_point = random.randint(1, self.n_intermediate - 1)
        child_path = parent1[:crossover_point] + parent2[crossover_point:]
        return child_path

    def mutate(self, path):
        """ Mutate a path by randomly changing one of its intermediate points """
        if random.random() < self.mutation_rate:
            idx = random.randint(1, self.n_intermediate)  # Only mutate intermediate points
            path[idx] = self.random_point_in_free_space()
        return path

    def evolve(self):
        """ Perform the evolution process across generations """
        for generation in range(self.generations):
            selected_population = self.selection()
            new_population = selected_population[:]

            # Crossover
            while len(new_population) < self.population_size:
                parent1 = random.choice(selected_population)
                parent2 = random.choice(selected_population)
                child = self.crossover(parent1, parent2)
                new_population.append(child)

            # Mutation
            new_population = [self.mutate(path) for path in new_population]

            self.population = new_population

            best_path = min(self.population, key=lambda path: self.fitness(path))
            if self.fitness(best_path) != float('inf'):
                print(f"Valid path found in generation {generation}")
                print("original path coordinates:", best_path)
                return best_path

        print("No valid path found within the specified generations")
        return min(self.population, key=lambda path: self.fitness(path))


class Plotting:
    def __init__(self, x_start, x_goal, obs_rectangle):
        self.xI, self.xG = x_start, x_goal
        self.obs_rectangle = obs_rectangle

    def plot(self, path, name):
        fig, ax = plt.subplots()
        ax.set_facecolor('white')

        # Plot obstacles
        for (ox, oy, w, h) in self.obs_rectangle:
            ax.add_patch(patches.Rectangle((ox, oy), w, h, edgecolor='black', facecolor='gray', fill=True))

        # Plot start and goal
        plt.plot(self.xI[0], self.xI[1], "bs", linewidth=3, label="Start")
        plt.plot(self.xG[0], self.xG[1], "gs", linewidth=3, label="Goal")

        # Plot path
        if path:
            plt.plot([x[0] for x in path], [x[1] for x in path], "r", linewidth=2, label="Path")

        plt.title(name)
        plt.axis("equal")
        plt.show()


def calculate_path_length(path):
    """ Calculate the total length of a path """
    total_length = 0
    for i in range(len(path) - 1):
        # 计算相邻点之间的欧几里得距离
        segment_length = np.linalg.norm(np.array(path[i + 1]) - np.array(path[i]))
        total_length += segment_length
    return total_length


def optimize_path(path, env):
    """
    Optimize the path by iterating from the goal to the start.
    Try to skip intermediate points if a straight line between two points has no obstacles.

    Parameters:
        path (list): The original path to be optimized.
        env (Env): The environment containing obstacle information (has is_collision_in_segment method).

    Returns:
        optimized_path (list): The optimized path from start to goal.
    """
    optimized_path = [path[-1]]  # Start from the goal
    current_point = path[-1]  # Begin optimization from the goal (last point in the path)

    # Start checking points from second to last (path[-2]) down to the first point (path[0])
    i = len(path) - 2

    while i >= 0:
        # Check if there is a direct path from the current point to the point at path[i]
        if not env.is_collision_in_segment(path[i], current_point):
            # If no collision, skip all points between current_point and path[i]
            i -= 1  # Try to connect to the next point
        else:
            # If collision occurs, keep the point and set it as the new current_point
            optimized_path.insert(0, path[i + 1])  # Keep the last non-colliding point
            current_point = path[i + 1]  # Update current_point
            i -= 1

    # Finally, add the start point (path[0])
    optimized_path.insert(0, path[0])

    return optimized_path


# Initialize environment and genetic algorithm
env = Env()
start = (2, 2)
goal = (845, 805)

start_time = time.time()

ga = GeneticAlgorithmPathPlanner(start, goal, env)
best_path = ga.evolve()
plotter = Plotting(start, goal, env.obs_rectangle)

best_path_length = calculate_path_length(best_path)
print(f"Total path length: {best_path_length}")

modified_path = optimize_path(best_path, env)
print("modified path is", modified_path)
modified_path_length = calculate_path_length(modified_path)
print(f"Total path length: {modified_path_length}")

end_time = time.time()

print("running time is", end_time - start_time)

plotter.plot(best_path, "original GA Path")
plotter.plot(modified_path, "modified GA path")


# save data into an Excel file
# excel_file_path = 'data.xlsx'
# try:
#     # 尝试打开现有的 Excel 文件
#     workbook = openpyxl.load_workbook(excel_file_path)
#     sheet = workbook.active
# except FileNotFoundError:
#     # 如果文件不存在，则创建一个新的工作簿
#     workbook = openpyxl.Workbook()
#     sheet = workbook.active
#
# def find_first_empty_row(sheet, column):
#     row = 1  # 行号从1开始
#     while sheet.cell(row=row, column=column).value is not None:
#         row += 1
#     return row
#
# # 找到第一列和第二列的第一个空行
# first_empty_row_a = find_first_empty_row(sheet, 1)  # 第一列 (a)
# first_empty_row_b = find_first_empty_row(sheet, 2)  # 第二列 (b)
#
# # 依次将列表 a 和 b 的值写入空行开始的位置
#
# sheet.cell(row=first_empty_row_a, column=1, value=modified_path_length)  # 写入第一列
# sheet.cell(row=first_empty_row_b, column=2, value=end_time - start_time)  # 写入第二列
#
# # 保存 Excel 文件
# workbook.save(excel_file_path)
#
# print(f"数据已保存到 {excel_file_path}")






